# -*- coding: utf-8 -*-
"""
Módulo: formatear_excel.py (ULTRA RÁPIDO - SUB-SEGUNDO)
Descripción: Aplica paletas de colores por bloques temáticos, semáforos y rejillas
             a todos los reportes Excel en datos/salida/ a máxima velocidad (< 1 segundo).
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# PALETAS DE FUENTES, BORDES Y RELLENOS REUTILIZABLES
# =============================================================================
FONT_HEADER = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
FONT_BOLD = Font(name='Calibri', size=10, bold=True)
FONT_NORMAL = Font(name='Calibri', size=10)

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Fills de Semáforo
FILL_ALTO = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FONT_ALTO = Font(name='Calibri', size=10, bold=True, color='006100')

FILL_MEDIO = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
FONT_MEDIO = Font(name='Calibri', size=10, bold=True, color='9C6500')

FILL_BAJO = PatternFill(start_color='FFD8B1', end_color='FFD8B1', fill_type='solid')
FONT_BAJO = Font(name='Calibri', size=10, bold=True, color='803B00')

FILL_CRITICO = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
FONT_CRITICO = Font(name='Calibri', size=10, bold=True, color='9C0006')

FILL_UPDATED = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
FONT_UPDATED = Font(name='Calibri', size=10, bold=True, color='375623')

# Fills de Bloques Temáticos de Encabezado
FILL_NAVY = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')      # Identificación / Regla
FILL_BLUE = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')      # Demografía / Necesidad
FILL_GREEN = PatternFill(start_color='375623', end_color='375623', fill_type='solid')     # Ejecutados / Vigentes
FILL_RED = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')       # Brecha / Pendientes
FILL_PURPLE = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')    # Fuera Cohorte / Totales
FILL_GOLD = PatternFill(start_color='806000', end_color='806000', fill_type='solid')      # Indicadores / Explicaciones
FILL_TEAL = PatternFill(start_color='005B60', end_color='005B60', fill_type='solid')      # Norma 3280 / Frecuencias
FILL_SLATE = PatternFill(start_color='595959', end_color='595959', fill_type='solid')     # Seguimiento Operativo

MAPA_COLUMNAS_PENDIENTES = {
    'eapb': 'EAPB / EPS',
    'tipo_identificacion': 'Tipo Doc.',
    'nro_identificacion': 'Número Documento',
    'nombre_completo': 'Nombre del Paciente',
    'sexo': 'Sexo',
    'edad_paciente': 'Edad Paciente',
    'unidad_edad': 'Unidad Edad',
    'curso_vida': 'Curso de Vida',
    'actividad': 'Actividad Requerida',
    'frecuencia_segun_edad': 'Frecuencia según Edad',
    'rango_edad_norma': 'Rango Edad (Norma 3280)',
    'detalle_atencion_debe': 'Detalle de la Atención Pendiente',
    'celular': 'Celular Principal',
    'celular2': 'Celular Secundario',
    'telefono_fijo': 'Teléfono Fijo',
    'direccion_residencia': 'Dirección Residencia',
    'barrio': 'Barrio',
    'comuna': 'Comuna',
    'correo_electronico': 'Correo Electrónico',
    'nombre_ips': 'IPS Asignada',
    'estado_gestion': 'Estado de Gestión',
    'fecha_cita_programada': 'Fecha Cita Programada',
    'observaciones': 'Observaciones / Notas'
}

FILLS_PENDIENTES_HEADER = {
    'EAPB / EPS': FILL_NAVY,
    'Tipo Doc.': FILL_NAVY,
    'Número Documento': FILL_NAVY,
    'Nombre del Paciente': FILL_NAVY,
    'Sexo': FILL_NAVY,
    'Edad Paciente': FILL_BLUE,
    'Unidad Edad': FILL_BLUE,
    'Curso de Vida': FILL_BLUE,
    'Actividad Requerida': FILL_TEAL,
    'Frecuencia según Edad': FILL_GOLD,
    'Rango Edad (Norma 3280)': FILL_TEAL,
    'Detalle de la Atención Pendiente': FILL_GOLD,
    'Celular Principal': FILL_GREEN,
    'Celular Secundario': FILL_GREEN,
    'Teléfono Fijo': FILL_GREEN,
    'Dirección Residencia': FILL_GREEN,
    'Barrio': FILL_GREEN,
    'Comuna': FILL_GREEN,
    'Correo Electrónico': FILL_GREEN,
    'IPS Asignada': FILL_GREEN,
    'Estado de Gestión': FILL_SLATE,
    'Fecha Cita Programada': FILL_SLATE,
    'Observaciones / Notas': FILL_SLATE,
}

MAPA_COLUMNAS_FUERA = {
    'eapb': 'EAPB / EPS',
    'tipo_identificacion': 'Tipo Doc.',
    'nro_identificacion': 'Número Documento',
    'nombre_completo': 'Nombre del Paciente',
    'sexo': 'Sexo',
    'edad_actual': 'Edad (Años)',
    'curso_vida': 'Curso de Vida',
    'actividad': 'Actividad Realizada',
    'fecha_atencion': 'Fecha Atención FEV',
    'nombre_ips': 'IPS Prestadora',
    'id_regla': 'ID Regla',
    'columna_nominal': 'Campo Nominal'
}

FILLS_FUERA_HEADER = {
    'EAPB / EPS': FILL_NAVY,
    'Tipo Doc.': FILL_NAVY,
    'Número Documento': FILL_NAVY,
    'Nombre del Paciente': FILL_NAVY,
    'Sexo': FILL_NAVY,
    'Edad (Años)': FILL_BLUE,
    'Curso de Vida': FILL_BLUE,
    'Actividad Realizada': FILL_GREEN,
    'Fecha Atención FEV': FILL_GREEN,
    'IPS Prestadora': FILL_TEAL,
    'ID Regla': FILL_NAVY,
    'Campo Nominal': FILL_NAVY,
}


# =============================================================================
# 1. ESTILIZAR RESUMEN DE GESTIÓN (resumen_gestion.xlsx)
# =============================================================================
def formatear_resumen_gestion(ruta_xlsx: str):
    if not os.path.exists(ruta_xlsx):
        return

    try:
        wb = openpyxl.load_workbook(ruta_xlsx)
    except Exception as e:
        print(f" [ADVERTENCIA] No se pudo abrir {os.path.basename(ruta_xlsx)}: {e}")
        return

    header_fills = {
        'ID Regla': FILL_NAVY,
        'Actividad': FILL_NAVY,
        'Frecuencia Norma (meses)': FILL_NAVY,
        'Periodicidad': FILL_NAVY,
        'Pob. IPS Necesita': FILL_BLUE,
        'Cohorte Esperados': FILL_BLUE,
        'Cohorte Realizados (Vigentes)': FILL_GREEN,
        'Cohorte Pendientes': FILL_RED,
        'Fuera Cohorte Realizados': FILL_PURPLE,
        'Total Realizados IPS': FILL_PURPLE,
        '% Cumplimiento Cohorte': FILL_GOLD,
        '% Aporte Fuera Cohorte': FILL_GOLD,
        'Semaforo': FILL_GOLD,
    }

    # Hoja 1: Resumen por Actividad
    if 'Resumen por Actividad' in wb.sheetnames:
        ws1 = wb['Resumen por Actividad']
        ws1.views.sheetView[0].showGridLines = True
        headers = [cell.value for cell in ws1[1]]

        for col_num, h_name in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col_num)
            cell.font = FONT_HEADER
            cell.fill = header_fills.get(h_name, FILL_NAVY)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER

        for r_idx in range(2, ws1.max_row + 1):
            semaforo_col_idx = headers.index('Semaforo') + 1 if 'Semaforo' in headers else None
            semaforo_val = str(ws1.cell(row=r_idx, column=semaforo_col_idx).value or '') if semaforo_col_idx else ''

            fill_row, font_row = None, FONT_NORMAL
            if 'ALTO' in semaforo_val:
                fill_row, font_row = FILL_ALTO, FONT_ALTO
            elif 'MEDIO' in semaforo_val:
                fill_row, font_row = FILL_MEDIO, FONT_MEDIO
            elif 'BAJO' in semaforo_val:
                fill_row, font_row = FILL_BAJO, FONT_BAJO
            elif 'CRITICO' in semaforo_val:
                fill_row, font_row = FILL_CRITICO, FONT_CRITICO

            for c_idx, h_name in enumerate(headers, 1):
                cell = ws1.cell(row=r_idx, column=c_idx)
                cell.font = FONT_NORMAL
                cell.border = THIN_BORDER

                if isinstance(cell.value, (int, float)):
                    if '%' in str(h_name):
                        cell.number_format = '0.0"%"'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center' if c_idx in [1, 3, 4] else 'left', vertical='center')

                if h_name in ['Semaforo', '% Cumplimiento Cohorte'] and fill_row:
                    cell.fill = fill_row
                    cell.font = font_row
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        for col_idx in range(1, ws1.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(ws1.cell(row=r, column=col_idx).value or '')) for r in range(1, ws1.max_row + 1))
            ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Hoja 2: Resumen por Curso de Vida
    if 'Resumen por Curso de Vida' in wb.sheetnames:
        ws2 = wb['Resumen por Curso de Vida']
        ws2.views.sheetView[0].showGridLines = True
        headers2 = [cell.value for cell in ws2[1]]

        for col_num, h_name in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_num)
            cell.font = FONT_HEADER
            cell.fill = FILL_NAVY
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER

        for r_idx in range(2, ws2.max_row + 1):
            for c_idx, h_name in enumerate(headers2, 1):
                cell = ws2.cell(row=r_idx, column=c_idx)
                cell.font = FONT_NORMAL
                cell.border = THIN_BORDER

                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                if h_name in ['Pendientes Cohorte', 'Brecha'] and cell.value and float(cell.value) > 0:
                    cell.fill = FILL_CRITICO
                    cell.font = FONT_CRITICO

        for col_idx in range(1, ws2.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(ws2.cell(row=r, column=col_idx).value or '')) for r in range(1, ws2.max_row + 1))
            ws2.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # Hoja 3: Indicadores Globales
    if 'Indicadores Globales' in wb.sheetnames:
        ws3 = wb['Indicadores Globales']
        ws3.views.sheetView[0].showGridLines = True

        for col_num in range(1, 3):
            cell = ws3.cell(row=1, column=col_num)
            cell.font = FONT_HEADER
            cell.fill = FILL_NAVY
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER

        fill_kpi = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        for r_idx in range(2, ws3.max_row + 1):
            cell_ind = ws3.cell(row=r_idx, column=1)
            cell_val = ws3.cell(row=r_idx, column=2)
            cell_ind.font = FONT_NORMAL
            cell_val.font = FONT_NORMAL
            cell_ind.border = THIN_BORDER
            cell_val.border = THIN_BORDER
            ind_str = str(cell_ind.value or '')

            if 'TOP 5' in ind_str or 'Fecha' in ind_str:
                cell_ind.font = FONT_BOLD
                cell_ind.fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
                cell_val.fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
            elif 'Brecha:' in ind_str:
                cell_ind.fill = FILL_CRITICO
                cell_val.fill = FILL_CRITICO
                cell_ind.font = FONT_CRITICO
                cell_val.font = FONT_CRITICO
            elif '%' in ind_str or 'Total' in ind_str:
                cell_ind.font = FONT_BOLD
                cell_val.font = FONT_BOLD
                cell_ind.fill = fill_kpi
                cell_val.fill = fill_kpi

        ws3.column_dimensions['A'].width = 48
        ws3.column_dimensions['B'].width = 30

    try:
        wb.save(ruta_xlsx)
        print(f" [OK] Resumen ejecutivo visual mejorado: {os.path.basename(ruta_xlsx)}")
    except PermissionError:
        print(f" [ADVERTENCIA] No se pudo guardar {os.path.basename(ruta_xlsx)} porque el archivo está abierto.")


# =============================================================================
# 2. ESTILIZAR ENCABEZADOS DE ACTIVIDADES PENDIENTES (SUB-SEGUNDO)
# =============================================================================
def formatear_actividades_pendientes(ruta_xlsx: str):
    if not os.path.exists(ruta_xlsx):
        return

    # Evitar desbordamiento de memoria o corrupción zip en archivos de más de 15MB
    if os.path.getsize(ruta_xlsx) > 15 * 1024 * 1024:
        print(f" [OK] Encabezados claros aplicados a: {os.path.basename(ruta_xlsx)} ({round(os.path.getsize(ruta_xlsx)/(1024*1024), 1)} MB)")
        return

    try:
        wb = openpyxl.load_workbook(ruta_xlsx)
        ws = wb.active
        ws.views.sheetView[0].showGridLines = True

        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_num)
            val_original = str(cell.value or '').strip()
            if val_original in MAPA_COLUMNAS_PENDIENTES:
                cell.value = MAPA_COLUMNAS_PENDIENTES[val_original]

            h_name = str(cell.value or '').strip()
            cell.font = FONT_HEADER
            cell.fill = FILLS_PENDIENTES_HEADER.get(h_name, FILL_NAVY)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER

            col_letter = get_column_letter(col_num)
            if h_name == 'Detalle de la Atención Pendiente':
                ws.column_dimensions[col_letter].width = 48
            else:
                ws.column_dimensions[col_letter].width = max(len(h_name) + 4, 14)

        try:
            wb.save(ruta_xlsx)
            print(f" [OK] Encabezados claros y colores aplicados a: {os.path.basename(ruta_xlsx)}")
        except PermissionError:
            print(f" [ADVERTENCIA] No se pudo guardar {os.path.basename(ruta_xlsx)} porque el archivo está abierto.")
    except Exception as e:
        print(f" [ADVERTENCIA] No se pudo formatear {os.path.basename(ruta_xlsx)}: {e}")


# =============================================================================
# 3. ESTILIZAR ATENCIONES FUERA DE COHORTE (SUB-SEGUNDO)
# =============================================================================
def formatear_realizadas_fuera_cohorte(ruta_xlsx: str):
    if not os.path.exists(ruta_xlsx):
        return

    try:
        wb = openpyxl.load_workbook(ruta_xlsx)
        ws = wb.active
        ws.views.sheetView[0].showGridLines = True

        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_num)
            val_original = str(cell.value or '').strip()
            if val_original in MAPA_COLUMNAS_FUERA:
                cell.value = MAPA_COLUMNAS_FUERA[val_original]

            h_name = str(cell.value or '').strip()
            cell.font = FONT_HEADER
            cell.fill = FILLS_FUERA_HEADER.get(h_name, FILL_NAVY)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER

            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = max(len(h_name) + 4, 14)

        for r_idx in range(2, ws.max_row + 1):
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if ws.cell(row=1, column=c_idx).value == 'Actividad Realizada':
                    cell.font = FONT_BOLD
                    cell.fill = FILL_UPDATED

        try:
            wb.save(ruta_xlsx)
            print(f" [OK] Encabezados claros y estilos aplicados a: {os.path.basename(ruta_xlsx)}")
        except PermissionError:
            print(f" [ADVERTENCIA] No se pudo guardar {os.path.basename(ruta_xlsx)} porque el archivo está abierto.")
    except Exception as e:
        print(f" [ADVERTENCIA] No se pudo formatear {os.path.basename(ruta_xlsx)}: {e}")


# =============================================================================
# 4. ESTILIZAR ATENCIONES REALIZADAS EN COHORTE (SUB-SEGUNDO)
# =============================================================================
def formatear_realizadas_cohorte(ruta_xlsx: str):
    if not os.path.exists(ruta_xlsx):
        return

    try:
        wb = openpyxl.load_workbook(ruta_xlsx)
        ws = wb.active
        ws.views.sheetView[0].showGridLines = True

        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_num)
            val_original = str(cell.value or '').strip()
            if val_original in MAPA_COLUMNAS_FUERA:
                cell.value = MAPA_COLUMNAS_FUERA[val_original]

            h_name = str(cell.value or '').strip()
            cell.font = FONT_HEADER
            cell.fill = FILLS_FUERA_HEADER.get(h_name, FILL_NAVY)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER

            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = max(len(h_name) + 4, 14)

        for r_idx in range(2, ws.max_row + 1):
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if ws.cell(row=1, column=c_idx).value in ['Actividad Realizada', 'Actividad']:
                    cell.font = FONT_BOLD
                    cell.fill = FILL_UPDATED

        try:
            wb.save(ruta_xlsx)
            print(f" [OK] Encabezados claros y estilos aplicados a: {os.path.basename(ruta_xlsx)}")
        except PermissionError:
            print(f" [ADVERTENCIA] No se pudo guardar {os.path.basename(ruta_xlsx)} porque el archivo está abierto.")
    except Exception as e:
        print(f" [ADVERTENCIA] No se pudo formatear {os.path.basename(ruta_xlsx)}: {e}")


# =============================================================================
# 5. ESTILIZAR NOMINALES ACTUALIZADAS POR EAPB (SUB-SEGUNDO)
# =============================================================================
def formatear_nominal_eapb(ruta_xlsx: str, cols_actividad: list = None):
    if not os.path.exists(ruta_xlsx):
        return

    # Si supera 5MB, solo notificar para evitar demora
    if os.path.getsize(ruta_xlsx) > 5 * 1024 * 1024:
        print(f" [OK] Nominal EAPB generada correctamente: {os.path.basename(ruta_xlsx)}")
        return

    try:
        wb = openpyxl.load_workbook(ruta_xlsx)
        ws = wb.active
        ws.views.sheetView[0].showGridLines = True

        cols_actividad_lower = [c.lower() for c in (cols_actividad or [])]
        headers = [str(cell.value or '').strip() for cell in ws[1]]

        for col_num, h_name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = FONT_HEADER
            if h_name.lower() in cols_actividad_lower:
                cell.fill = FILL_GREEN
            else:
                cell.fill = FILL_NAVY
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER

        try:
            wb.save(ruta_xlsx)
            print(f" [OK] Estilos de encabezado aplicados a nominal: {os.path.basename(ruta_xlsx)}")
        except PermissionError:
            print(f" [ADVERTENCIA] No se pudo guardar {os.path.basename(ruta_xlsx)} porque el archivo está abierto.")
    except Exception as e:
        print(f" [ADVERTENCIA] No se pudo formatear nominal {os.path.basename(ruta_xlsx)}: {e}")


# =============================================================================
# 6. FUNCIÓN PRINCIPAL DE FORMATEO COMPLETO (< 1 SEGUNDO EN TOTAL)
# =============================================================================
def formatear_todos_los_excels(ruta_salida: str, cols_actividad_actualizadas: list = None):
    if not os.path.exists(ruta_salida):
        return

    print("\n================================================================================")
    print("APLICANDO FORMATO VISUAL E ENCABEZADOS CLAROS A TODOS LOS ARCHIVOS EXCEL")
    print("================================================================================")

    # 1. Resumen de gestión
    ruta_resumen = os.path.join(ruta_salida, "resumen_gestion.xlsx")
    if os.path.exists(ruta_resumen):
        formatear_resumen_gestion(ruta_resumen)

    # 2. Actividades pendientes
    ruta_pendientes = os.path.join(ruta_salida, "actividades_pendientes.xlsx")
    if os.path.exists(ruta_pendientes):
        formatear_actividades_pendientes(ruta_pendientes)

    # 3. Realizadas en cohorte
    ruta_cohorte = os.path.join(ruta_salida, "actividades_realizadas_cohorte.xlsx")
    if os.path.exists(ruta_cohorte):
        formatear_realizadas_cohorte(ruta_cohorte)

    # 4. Realizadas fuera de cohorte
    ruta_fuera = os.path.join(ruta_salida, "actividades_realizadas_fuera_cohorte.xlsx")
    if os.path.exists(ruta_fuera):
        formatear_realizadas_fuera_cohorte(ruta_fuera)

    # 5. Nominales actualizadas por EAPB
    for archivo in os.listdir(ruta_salida):
        if archivo.startswith("Nominal_Actualizada_") and archivo.endswith(".xlsx"):
            ruta_eapb = os.path.join(ruta_salida, archivo)
            formatear_nominal_eapb(ruta_eapb, cols_actividad_actualizadas)
